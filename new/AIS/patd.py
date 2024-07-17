import re
import openpyxl
from openpyxl.styles import PatternFill
 
def chose(file,name):
    dwb = openpyxl.load_workbook(file) #讀取檔案
    s1 = dwb['all']  #讀取總表
    wb = openpyxl.Workbook()#開新的工作表

    
    data =[[0]*65 for t in range(26)]#建立空串列
    result=[[0]*65 for t in range(26)]#建立空串列

    for x in range(0,s1.max_row):#讀資料
                for y in range(0,s1.max_column):
                    data[x][y]=s1.cell(x+1,y+1).value
    for x in range(0,s1.max_row):#讀資料
                for y in range(0,s1.max_column):
                    result[x][y]=s1.cell(x+1,y+1).value
    yave=130#年平均
    total=0#初始化計數器
    count=0#初始化計數器

    for x in range(2,len(data)):#計算月平均
        for y in range(1,len(data[0])):
            if data[x][y]!=0:     
                total=total+data[x][y]
                count=count+1
    mave=total/count

    if yave<=mave:#比較年平均跟月平均差異，並開始執行演算法
        for y in range(1,len(data[0])):
            for x in range(2,len(data)):
                if x<3:
                    if (data[x][y]+data[x+1][y])/2 >=mave:
                        result[x][y]=1
                    else:
                        result[x][y]=0
                elif x==len(data)-1:
                    if (data[x][y]+data[x-1][y])/2 >=mave:
                        result[x][y]=1
                    else:
                        result[x][y]=0
                else:
                    if (data[x][y]+data[x+1][y]+data[x-1][y])/3 >=mave:
                        result[x][y]=1
                    else:
                        result[x][y]=0
    else:
        for y in range(1,len(data[0])):
            for x in range(2,len(data)):
                if x<3:
                    if (data[x][y]+data[x+1][y])/2 >=mave:
                        result[x][y]=1
                    else:
                        result[x][y]=0
                elif x==len(data)-1:
                    if (data[x][y]+data[x-1][y])/2 >=mave:
                        result[x][y]=1
                    else:
                        result[x][y]=0
                else:
                    if (data[x][y]+data[x+1][y]+data[x-1][y])/3 >=mave:
                        result[x][y]=1
                    else:
                        result[x][y]=0

    sheet = wb.create_sheet("patd")#建立PATD試算表
    fill=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")#建立填充顏色資料
    for x in range(0,len(result)):#填入表格
                for y in range(0,len(result[0])):
                    sheet.cell(x+1,y+1).value=result[x][y]
                    
    for x in range(2,sheet.max_row):#填充顏色
                for y in range(1,sheet.max_column):
                    if sheet.cell(x+1,y+1).value==1:
                        sheet.cell(x+1,y+1).fill=fill

    sheet2 = wb.create_sheet("patd2")#建立不同格式PATD試算表
    for x in range(0,sheet.max_row):#填填入資料
        for y in range(0,sheet.max_column):
            sheet2.cell(y+2,x+1).value=sheet.cell(x+1,y+1).value
    for x in range(2,sheet2.max_row):#填充顏色
        for y in range(2,sheet2.max_column):
            if sheet2.cell(x+1,y+1).value==1:
                sheet2.cell(x+1,y+1).fill=fill
    for x in range(0,sheet2.max_row):#調整格式
        for col in sheet2.columns:
                col_name = re.findall('\w\d', str(col[0]))
                col_name = col_name[1]
                col_name = re.findall('\w', str(col_name))[0]
                sheet2.column_dimensions[col_name].width =9
                sheet2.row_dimensions[x+2].height =50
    sheet2.cell(1,1).value="月平均值"#填入月平均
    sheet2.cell(1,2).value=mave#填入月平均
    del wb["Sheet"]#刪除空試算表

    wb.save(name)#儲存檔案
    print("done")
    return(name)