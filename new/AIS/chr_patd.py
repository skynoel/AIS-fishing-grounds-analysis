import openpyxl
from openpyxl.styles import PatternFill

def chr(file,name):
    dwb = openpyxl.load_workbook(file)  #讀取檔案
    s1 = dwb['patd2']  #讀取PATD表
    wb = openpyxl.Workbook()#開新的工作表


    data =[[0]*26 for t in range(66)]#建立空串列
    result =[[0]*26 for t in range(66)]#建立空串列

    for x in range(0,s1.max_row):#讀資料
                for y in range(0,s1.max_column):
                    data[x][y]=s1.cell(x+1,y+1).value
    for x in range(0,s1.max_row):#讀資料
                for y in range(0,s1.max_column):
                    result[x][y]=s1.cell(x+1,y+1).value

    for x in range(2,len(data)):#琪何規則
        for y in range(4,len(data[0])-2):
            if (data[x][y+2]+data[x][y+1]+data[x][y]+data[x][y-1]+data[x][y-2])>=3:
                    if result[x][y] !=1:
                        result[x][y]=2
    for x in range(2,len(data)):#琪何規則
        for y in range(2,len(data[0])):
            if (data[x][1]=="A"):
                if result[x][y] ==1 or result[x][y]==2:
                        if result[x+1][y] !=1:
                            result[x+1][y]=2
            if (data[x][1]=="B"):
                if result[x][y] ==1 or result[x][y]==2:
                        if result[x-1][y] !=1:
                            result[x-1][y]=2

    sheet = wb.create_sheet("chr_patd")#建立CHRPATD試算表
    for x in range(0,len(result)):#填入資料
                for y in range(0,len(result[0])):
                    sheet.cell(x+1,y+1).value=result[x][y]


    fill=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")#建立顏色資訊
    fill2=PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")#建立顏色資訊

    for x in range(2,sheet.max_row):#上色
                for y in range(1,sheet.max_column):
                    if sheet.cell(x+1,y+1).value==1:
                        sheet.cell(x+1,y+1).fill=fill
                    if sheet.cell(x+1,y+1).value==2:
                        sheet.cell(x+1,y+1).fill=fill2
    del wb["Sheet"]#刪除空試算表
    wb.save(name)#儲存檔案
    print("done")
    return name