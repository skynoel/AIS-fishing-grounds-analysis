import numpy
import openpyxl
import pandas as pd
import os
from tqdm import tqdm
import calendar
import matplotlib.pyplot as plt

def avperm(folder_path,name,year,month): #將經過時間碼標記的原始資料轉變為個小時每分鐘平均封包數的資料，並回傳檔名
    wb = openpyxl.Workbook()#建立試算表

    count2=[[0]*65 for t in range(26)]#建立用於總表的空串列

    folders = os.listdir(folder_path)#列出目標資料夾中所有內容
    for folder in tqdm(folders):#遍歷目標資料夾所有的所有內容
        folder_path1 = os.path.join(folder_path, folder)#取出目標資料夾中內容的路徑
        if os.path.isdir(folder_path1):#判斷該路徑是否為資料夾
            count = [[0]*65 for t in range(4)]#建立用於每日表格的空串列
            count[0][0]="RMC_DATE_YEAR"#初始化每日表格
            count[0][1]="RMC_DATE_MON"#初始化每日表格
            count[0][2]="RMC_DATE_DAY"#初始化每日表格
            count[0][3]="RMC_TAKEN_AT_HOUR"#初始化每日表格
            count[0][4]="CHANNEL"#初始化每日表格
            for y in range(5):#初始化每日表格
                count[1][y]=" "
            for y in range(60):#初始化每日表格
                count[1][y+5]=y
            for y in range(60):#初始化每日表格
                count[0][y+5]=" "
            row=2#初始化計數器
            rowchange=30#初始化計數器
            r2=40#初始化計數器
            r3=13
            items = os.listdir(folder_path1)#列出資料夾中所有內容
            for item in tqdm(items):#遍歷資料夾所有的所有內容
                item_path = os.path.join(folder_path1, item)#取出資料夾中內容的路徑
                if os.path.isfile(item_path):#判斷該路徑是否為檔案
                    print("檔案:", item)#輸出檔案名稱
                    df=pd.read_excel(item_path)#讀取檔案
                    for i in tqdm(range(df.shape[0])):#檔案內容迴圈
                        data=df.iloc[i]#取出第i行資料
                        #a=isinstance(data.iloc[3],numpy.int64)#判斷正確格式
                        #b=isinstance(data.iloc[2],numpy.int64)#判斷正確格式
                        if isinstance(data.iloc[3],str)==False and pd.isnull(data.iloc[3])==False :#如果兩個格式都正確就開始處理(下面為核心算法)                            
                            #print(type(data.iloc[3]))
                            if rowchange<int(data.iloc[3]):
                                row=row+2
                                count.append([0]*65)
                                count.append([0]*65)
                                rowchange=int(data.iloc[3])
                                for k in range(0,4):
                                    count[row][k]=int(data.iloc[k])
                                    count[row+1][k]=int(data.iloc[k])
                            elif r2<int(data.iloc[2]):
                                row=row+2
                                count.append([0]*65)
                                count.append([0]*65)
                                r2=int(data.iloc[2])
                                for k in range(0,4):
                                    count[row][k]=int(data.iloc[k])
                                    count[row+1][k]=int(data.iloc[k])
                            elif r3<int(data.iloc[1]):
                                row=row+2
                                count.append([0]*65)
                                count.append([0]*65)
                                r3=int(data.iloc[1])
                                for k in range(0,4):
                                    count[row][k]=int(data.iloc[k])
                                    count[row+1][k]=int(data.iloc[k])   
                            if data.iloc[8] == "A":
                                count[row][4]="A"
                                count[row][int(data.iloc[4])+5]=int(count[row][int(data.iloc[4])+5])+1
                            elif data.iloc[8] == "B":
                                count[row+1][4]="B"
                                count[row+1][int(data.iloc[4])+5]=int(count[row+1][int(data.iloc[4])+5])+1
                            rowchange=int(data.iloc[3])
                            r2=int(data.iloc[2])
                            r3=int(data.iloc[1])
                            if row==2:
                                for k in range(0,4):
                                    count[row][k]=int(data.iloc[k])
                                    count[row+1][k]=int(data.iloc[k])
                        elif isinstance(data.iloc[3],str)==True:
                            if data.iloc[3].isdigit():
                                if rowchange<int(data.iloc[3]):
                                    row=row+2
                                    count.append([0]*65)
                                    count.append([0]*65)
                                    rowchange=int(data.iloc[3])
                                    for k in range(0,4):
                                        count[row][k]=int(data.iloc[k])
                                        count[row+1][k]=int(data.iloc[k])
                                elif r2<int(data.iloc[2]):
                                    row=row+2
                                    count.append([0]*65)
                                    count.append([0]*65)
                                    r2=int(data.iloc[2])
                                    for k in range(0,4):
                                        count[row][k]=int(data.iloc[k])
                                        count[row+1][k]=int(data.iloc[k])
                                elif r3<int(data.iloc[1]):
                                    row=row+2
                                    count.append([0]*65)
                                    count.append([0]*65)
                                    r3=int(data.iloc[1])
                                    for k in range(0,4):
                                        count[row][k]=int(data.iloc[k])
                                        count[row+1][k]=int(data.iloc[k])
                                if data.iloc[8] == "A":
                                    count[row][4]="A"
                                    count[row][int(data.iloc[4])+5]=int(count[row][int(data.iloc[4])+5])+1
                                elif data.iloc[8] == "B":
                                    count[row+1][4]="B"
                                    count[row+1][int(data.iloc[4])+5]=int(count[row+1][int(data.iloc[4])+5])+1
                                rowchange=int(data.iloc[3])
                                r2=int(data.iloc[2])
                                r3=int(data.iloc[1])
                                if row==2:
                                    for k in range(0,4):
                                        count[row][k]=int(data.iloc[k])
                                        count[row+1][k]=int(data.iloc[k])
                        if count[row-2][3]==count[row][3] and count[row][3]==23 :
                            count[row][3]=0
                            count[row+1][3]=0
            sheetname=str(int(data.iloc[0]))+str(int(data.iloc[1]))+str(int(data.iloc[2]))#製造每日表格的名字
            sheet = wb.create_sheet(sheetname)#用上述名字建立表格

            for x in range(0,len(count)):#填充表格
                for y in range(0,len(count[0])):
                    sheet.cell(x+1,y+1).value=count[x][y]
            num1=0#初始化計數器
            num2=0#初始化計數器
            for x in range(2,len(count)):#計算平均值
                for y in range(5,len(count[0])):
                    if count[x][y]!= 0:
                        num1=num1+count[x][y]
                        num2=num2+1
                if num2!=0:
                    ave=num1/num2
                    day=count[x][2]*2-1
                    if count[x][4]=="A":
                        count2[count[x][3]+2][day]=ave
                    elif count[x][4]=="B":
                        count2[count[x][3]+2][day+1]=ave

    for i in range(2,26):#初始化總表
        count2[i][0]=i-2            
    x, y = calendar.monthrange(int(year), int(month))#日期設定
    k=1
    for j in range(1,y+1):#初始化總表
        count2[0][k]=str(month)+"/"+str(j)
        count2[0][k+1]=str(month)+"/"+str(j)
        k=k+2
    count2[0][0]="Date"
    count2[1][0]="channal"
    for c in range(1,65):#初始化總表
        if c%2==0:
            count2[1][c]="B"
        else:
            count2[1][c]="A"

    sheet = wb.create_sheet("all")#建立代表總表的試算表
    for x in range(0,len(count2)):#填入總表
                for y in range(0,len(count2[0])):
                    sheet.cell(x+1,y+1).value=count2[x][y]

    wb.save(name)  #儲存檔案
    print("done")
    return name


def graph(file):#繪圖部分
    data=openpyxl.load_workbook(file)#讀取檔案
    sheet=data['all']
    x=[]#初始化座標
    y=[]#初始化座標

    for k in range(2,sheet.max_column):#建立座標
        for i in range(3,sheet.max_row):
            A=sheet.cell(i,1).value
            x.append(A)
        for j in range(3,sheet.max_row):
            B=sheet.cell(j,k).value
            y.append(B)
        plt.plot(x, y, marker='.',color='red',linewidth=0.1,label=str(sheet.cell(1,k).value)+sheet.cell(2,k).value)
    plt.xlabel('hour')
    plt.ylabel('mad')
    plt.title('madfile')
    plt.savefig(file.rstrip(".xlsx")+".png") 
    plt.show()
    print("done")